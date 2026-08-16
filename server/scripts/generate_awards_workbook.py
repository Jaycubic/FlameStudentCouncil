# scripts/generate_awards_workbook.py
#
# Usage:
#   echo <json_data_base64> | python3 generate_awards_workbook.py \
#       <master_access_token> <master_refresh_token> <folder_id>
#
# json_data_base64: base64-encoded JSON read from STDIN (not a CLI arg).
# Passing it via stdin avoids the Linux ARG_MAX (E2BIG) limit.
#   Keys: all[]  byPosition{}
#   Each row includes: photo_url, student_id, name, email, gender, batch,
#   mobile_number, position_selected, academic_score, sports_score, cultural_score,
#   sports_verified_score, cultural_verified_score, academic_verified_score,
#   total_verified_score, submission_date, Workbook Link, Attachment
#
# Photo column logic:
#   • Column A: =IMAGE() formula using local server URL (https://flamestudentcouncil.in/api/photos/<student_id>)
#   • If photo_url is empty — cell is left blank (no broken image)
#
# Returns: { success, sheet_id, url }

import sys, json, os, time, base64, socket, tempfile, re
from datetime import datetime

from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build
from googleapiclient.errors import HttpError
from googleapiclient.http import MediaFileUpload

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"success": False, "error": "openpyxl not installed"}))
    sys.exit(0)

socket.setdefaulttimeout(120)

SCOPES = [
    'https://www.googleapis.com/auth/drive',
    'https://www.googleapis.com/auth/spreadsheets',
]

PROTECTED_COLS  = ['student_id', 'name', 'email']   # warning-only protection
VERIFIED_COLS   = ['sports_verified_score', 'cultural_verified_score', 'academic_verified_score', 'total_verified_score']
PHOTO_COL_WIDTH  = 22       # column A width (pixels/chars) for the photo column
PHOTO_ROW_HEIGHT = 90      # row height when photo is present

# Columns whose values are URLs — rendered as clickable hyperlinks with short labels
HYPERLINK_COLS = {
    'Workbook Link': 'Student Workbook',
    'Attachment':    'View PDF',
}

# Regex to strip illegal characters for openpyxl / XML export
ILLEGAL_CHAR_REGEX = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

# ─── Column definitions (data columns after the Photo column) ─────────────────
ELECTION_COLS = [
    'student_id', 'name', 'email', 'gender', 'batch', 'mobile_number',
    'position_selected', 'community_service', 'statement_of_purpose', 'more_info',
    'read_handbook', 'not_on_probation', 'tru_statement',
    'academic_score', 'sports_score', 'cultural_score',
    'sports_director_score', 'cultural_director_score',
    'sports_verified_score', 'cultural_verified_score', 'academic_verified_score',
    'total_verified_score', 'submission_date',
    'Workbook Link', 'Attachment'
]

# Styles
HEADER_FILL     = PatternFill("solid", fgColor="1E3A8A")   # dark blue  — data cols
HEADER_FONT     = Font(bold=True, color="FFFFFF", size=10)
LOCK_FILL       = PatternFill("solid", fgColor="FFF3CD")   # yellow tint — protected data cols
PHOTO_FILL      = PatternFill("solid", fgColor="0D9488")   # teal       — photo header
KEY_COL_FILL    = PatternFill("solid", fgColor="7C3AED")   # purple     — student_id/name/email header
GREY_HEADER_FILL= PatternFill("solid", fgColor="4B5563")   # dark grey  — verified cols header
GREY_DATA_FILL  = PatternFill("solid", fgColor="E5E7EB")   # light grey — verified cols data cells


# ─── Helpers ──────────────────────────────────────────────────────────────────

def execute_with_retry(request, max_retries=4):
    for attempt in range(max_retries):
        try:
            return request.execute()
        except HttpError as e:
            if e.resp.status in (429, 500, 503) and attempt < max_retries - 1:
                time.sleep(2 ** attempt)
            else:
                raise


def build_credentials(access_token, refresh_token):
    creds = Credentials(
        token=access_token,
        refresh_token=refresh_token,
        token_uri="https://oauth2.googleapis.com/token",
        client_id=os.environ.get('GOOGLE_CLIENT_ID'),
        client_secret=os.environ.get('GOOGLE_CLIENT_SECRET'),
        scopes=SCOPES,
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
    return creds


def get_photo_formula(photo_url):
    """Return =IMAGE() formula using the given URL, or '' if no URL."""
    if not photo_url:
        return ''
    return f'=IMAGE("{photo_url}")'


def sanitize_string(val):
    """Remove control characters that break openpyxl XML export."""
    if isinstance(val, str):
        return ILLEGAL_CHAR_REGEX.sub('', val)
    return val


def row_to_cells(record, cols, ri):
    """Extract values for the given column list from a record dict."""
    cells = []
    for col in cols:
        # Dynamic Excel formulas for verified scores (matching scaleScore piecewise formula)
        if col == 'sports_verified_score':
            val = f'=IFERROR(IF(VALUE(P{ri})<=150, ROUND(VALUE(P{ri})/15, 2), ROUND(MIN(10 + 0.05*(VALUE(P{ri})-150), 12), 2)), 0)'
        elif col == 'cultural_verified_score':
            val = f'=IFERROR(IF(VALUE(Q{ri})<=150, ROUND(VALUE(Q{ri})/15, 2), ROUND(MIN(10 + 0.05*(VALUE(Q{ri})-150), 12), 2)), 0)'
        elif col == 'academic_verified_score':
            val = f'=IFERROR(VALUE(O{ri}), 0)'
        elif col == 'total_verified_score':
            val = f'=MIN(30, ROUND(SUM(T{ri}:V{ri}) + IFERROR(VALUE(R{ri}), 0) + IFERROR(VALUE(S{ri}), 0), 2))'
        else:
            val = record.get(col, '')
            if isinstance(val, bool):
                val = 'True' if val else 'False'
            elif str(val).lower() == 'true':
                val = 'True'
            elif str(val).lower() == 'false':
                val = 'False'
            elif val is None:
                val = ''

            if col == 'submission_date' and val:
                try:
                    val = datetime.fromisoformat(str(val)).strftime('%Y-%m-%d')
                except Exception:
                    pass

            # Wrap URL columns in =HYPERLINK() so cell shows a short label
            if col in HYPERLINK_COLS and val:
                label = HYPERLINK_COLS[col]
                val = f'=HYPERLINK("{val}", "{label}")'
            elif col == 'Workbook Link' and not val:
                val = 'Matrix Not Opened'
            else:
                val = sanitize_string(str(val))
                # If a plain text field starts with '=' or '+' or '-' (e.g., bullet points), prefix with single quote
                if val.startswith(('=', '+', '-')) and col not in VERIFIED_COLS and col not in HYPERLINK_COLS:
                    val = f"'{val}"

        cells.append(val)
    return cells


def _thin_border():
    thin = Side(style='thin', color='D1D5DB')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


# ─── Sheet writer ─────────────────────────────────────────────────────────────

def write_sheet(ws, cols, rows):
    """
    Write a complete tab into the openpyxl worksheet.

    Column layout:
        A  → Photo  (=IMAGE formula, teal header)
        B+ → normal data columns (cols list)
    """
    border = _thin_border()

    # ── Header row ────────────────────────────────────────────────────────────
    ph = ws.cell(row=1, column=1, value='Photo')
    ph.font      = HEADER_FONT
    ph.fill      = PHOTO_FILL
    ph.alignment = Alignment(horizontal='center', vertical='center')
    ph.border    = border
    ws.column_dimensions['A'].width = PHOTO_COL_WIDTH

    for ci, col in enumerate(cols, 2):
        cell = ws.cell(row=1, column=ci, value=col)
        cell.font      = HEADER_FONT
        if col in PROTECTED_COLS:
            cell.fill = KEY_COL_FILL
        elif col in VERIFIED_COLS:
            cell.fill = GREY_HEADER_FILL
        else:
            cell.fill = HEADER_FILL

        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border    = border
        ws.column_dimensions[get_column_letter(ci)].width = max(len(col) + 4, 16)

    ws.freeze_panes = 'B2'   # freeze both Photo col and header row
    ws.row_dimensions[1].height = 32

    # ── Data rows ─────────────────────────────────────────────────────────────
    for ri, record in enumerate(rows, 2):
        photo_url = record.get('photo_url', '')
        formula   = get_photo_formula(photo_url)

        photo_cell = ws.cell(row=ri, column=1, value=formula)
        photo_cell.alignment = Alignment(horizontal='center', vertical='center')
        photo_cell.border    = border
        if photo_url:
            ws.row_dimensions[ri].height = PHOTO_ROW_HEIGHT

        cells = row_to_cells(record, cols, ri)
        for ci, val in enumerate(cells, 2):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border    = border
            cell.alignment = Alignment(vertical='center')
            col_name = cols[ci - 2]
            if col_name in PROTECTED_COLS:
                cell.fill = LOCK_FILL
            elif col_name in VERIFIED_COLS:
                cell.fill = GREY_DATA_FILL


# ─── Sheet protection ─────────────────────────────────────────────────────────

def protect_sheet_columns(sheets_service, spreadsheet_id, sheet_gid):
    """Warning-only protection on columns A–D and T–W."""
    body = {
        "requests": [
            {
                "addProtectedRange": {
                    "protectedRange": {
                        "range": {
                            "sheetId":           sheet_gid,
                            "startColumnIndex":  0,   # col A (Photo)
                            "endColumnIndex":    4,   # cols A-D exclusive (Photo,student_id,name,email)
                        },
                        "description": (
                            "⚠️ Photo / student_id / name / email are KEY fields. "
                            "Changes here will NOT be synced and may corrupt the local database."
                        ),
                        "warningOnly": True,
                    }
                }
            },
            {
                "addProtectedRange": {
                    "protectedRange": {
                        "range": {
                            "sheetId":           sheet_gid,
                            "startColumnIndex":  19,  # col T (sports_verified_score)
                            "endColumnIndex":    23,  # cols T-W exclusive (total_verified_score)
                        },
                        "description": (
                            "⚠️ Verified score columns are calculated dynamically via formulas. "
                            "Direct edits here may break automatic formula updates."
                        ),
                        "warningOnly": True,
                    }
                }
            }
        ]
    }
    try:
        execute_with_retry(
            sheets_service.spreadsheets().batchUpdate(
                spreadsheetId=spreadsheet_id,
                body=body
            )
        )
    except Exception as e:
        print(f"[WARN] Could not add sheet protection: {e}", file=sys.stderr)


# ─── Main ─────────────────────────────────────────────────────────────────────

def main():
    tmp_path = None
    try:
        if len(sys.argv) < 4:
            print(json.dumps({
                "success": False,
                "error": "Usage: master_access_token master_refresh_token folder_id (json_data_base64 via stdin)"
            }))
            return

        master_access_token  = sys.argv[1]
        master_refresh_token = sys.argv[2]
        folder_id            = sys.argv[3]

        try:
            json_data_b64 = sys.stdin.read().strip()
        except Exception as e:
            print(json.dumps({"success": False, "error": f"Failed to read stdin: {e}"}))
            return

        if not json_data_b64:
            print(json.dumps({"success": False, "error": "No data received on stdin"}))
            return

        try:
            data = json.loads(base64.b64decode(json_data_b64).decode('utf-8'))
        except Exception as e:
            print(json.dumps({"success": False, "error": f"Invalid JSON data: {e}"}))
            return

        all_rows    = data.get('all', [])
        by_position = data.get('byPosition', {})

        def safe_score(row, key='total_verified_score'):
            if not isinstance(row, dict):
                return 0.0
            try:
                val = row.get(key)
                if val is None or str(val).strip() == '' or str(val).strip() == '—':
                    return 0.0
                return float(val)
            except (ValueError, TypeError):
                return 0.0

        all_rows.sort(key=lambda r: safe_score(r), reverse=True)
        for pos in by_position:
            if isinstance(by_position[pos], list):
                by_position[pos].sort(key=lambda r: safe_score(r), reverse=True)

        try:
            master_creds = build_credentials(master_access_token, master_refresh_token)
        except Exception as e:
            print(json.dumps({"success": False, "error": f"Token error: {e}"}))
            return

        # Build workbook in memory
        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default blank sheet

        # Tab definitions: (tab_name, data_cols, rows)
        tabs = [('All Responses', ELECTION_COLS, all_rows)]

        tab_rows_map = {}
        for tab_name, cols, rows in tabs:
            ws = wb.create_sheet(tab_name)
            write_sheet(ws, cols, rows)
            tab_rows_map[tab_name] = rows

        # Save to temp file, then upload as Google Sheets
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        tmp_path = tmp.name
        tmp.close()
        wb.save(tmp_path)

        drive_service  = build('drive',  'v3', credentials=master_creds)
        sheets_service = build('sheets', 'v4', credentials=master_creds)

        file_metadata = {
            'name':     'FLAME Student Council — Election Master Workbook',
            'mimeType': 'application/vnd.google-apps.spreadsheet',
            'parents':  [folder_id],
        }
        media = MediaFileUpload(
            tmp_path,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            resumable=True,
        )
        uploaded = execute_with_retry(
            drive_service.files().create(
                body=file_metadata,
                media_body=media,
                fields='id,webViewLink',
            )
        )
        spreadsheet_id = uploaded['id']
        url            = uploaded['webViewLink']

        # Apply warning-only protection to each tab
        try:
            meta = execute_with_retry(sheets_service.spreadsheets().get(spreadsheetId=spreadsheet_id))
            gid_map = {s['properties']['title']: s['properties']['sheetId'] for s in meta.get('sheets', [])}
            for gid in gid_map.values():
                protect_sheet_columns(sheets_service, spreadsheet_id, gid)
        except Exception as protect_err:
            print(f"[WARN] Could not retrieve sheet metadata for protection: {protect_err}", file=sys.stderr)

        # ─── Post-upload: rewrite Photo column (col A) via Sheets API ─────────
        # Match photos by student_id (col B) to avoid order-mismatch bugs.
        # Build student_id → photo_url map from our in-memory data.
        sid_photo_maps = {}
        for tab_name, rows in tab_rows_map.items():
            photo_map = {}
            for record in rows:
                sid = str(record.get('student_id', '')).strip()
                if sid:
                    photo_map[sid] = record.get('photo_url', '')
            sid_photo_maps[tab_name] = photo_map

        # Read student_id column (B) from each tab in the uploaded spreadsheet
        read_ranges = [f"'{tn}'!B:B" for tn in tab_rows_map.keys()]
        try:
            batch_result = execute_with_retry(
                sheets_service.spreadsheets().values().batchGet(
                    spreadsheetId=spreadsheet_id,
                    ranges=read_ranges,
                )
            )
            range_results = batch_result.get('valueRanges', [])
        except Exception as _read_err:
            print(f"[WARN] Could not read student_id column for photo matching: {_read_err}", file=sys.stderr)
            range_results = []

        photo_data = []
        for i, tab_name in enumerate(tab_rows_map.keys()):
            photo_map = sid_photo_maps.get(tab_name, {})
            if not photo_map:
                continue

            if i < len(range_results):
                col_b_values = range_results[i].get('values', [])
            else:
                col_b_values = []

            if len(col_b_values) < 2:
                continue

            values = [['Photo']]   # header
            for row_data in col_b_values[1:]:
                sid = str(row_data[0]).strip() if row_data else ''
                photo_url = photo_map.get(sid, '')
                formula = get_photo_formula(photo_url)
                values.append([formula])
            photo_data.append({'range': f"'{tab_name}'!A1", 'values': values})

        if photo_data:
            try:
                execute_with_retry(
                    sheets_service.spreadsheets().values().batchUpdate(
                        spreadsheetId=spreadsheet_id,
                        body={'valueInputOption': 'USER_ENTERED', 'data': photo_data}
                    )
                )
            except Exception as _e:
                print(f"[WARN] Photo formula rewrite failed: {_e}", file=sys.stderr)

        print(json.dumps({"success": True, "sheet_id": spreadsheet_id, "url": url}))

    except HttpError as e:
        print(json.dumps({"success": False, "error": f"Google API error: {e}"}))
    except Exception as e:
        print(json.dumps({"success": False, "error": str(e)}))
    finally:
        if tmp_path and os.path.exists(tmp_path):
            try:
                os.unlink(tmp_path)
            except Exception:
                pass


if __name__ == '__main__':
    main()
